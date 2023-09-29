from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import os
import time

#Faz com que o codigo rode no Navegador Chrome
driver = webdriver.Chrome()
driver.get('https://app.localo.com/paywall')
driver.set_window_size(1920, 1080)

# Digita o email
email = 'seu-email'
campo_email = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/div/form/div[1]/input')
campo_email.send_keys(email)

# Digita a senha
senha = 'sua-senha'
campo_senha = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/div/form/div[2]/span/input')
campo_senha.send_keys(senha)

# Clicar em conectar
botao_conectar = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div[2]/div/form/button')
botao_conectar.click()
time.sleep(20)

workbook = load_workbook('dados.xlsx')
sheet = workbook['Plan1']
ultima_linha = sheet.max_row #Encontre a última linha preenchida na coluna A
for linha in range(2, ultima_linha + 1):
    # Ler o valor da célula na coluna A da linha atual
    valor = sheet[f'A{linha}'].value      
    campo_pesquisa = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/span/input') 
    campo_pesquisa.send_keys(valor)
    campo_pesquisa.send_keys(Keys.RETURN)
    time.sleep(20)

#Clique no resultado
resultado = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]')
resultado.click()
time.sleep(20)        

# Pesquisar Palavra Chave
palavra_chave = sheet['B2'].value
campo_palavra_chave = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/div/div[1]/span/input')
campo_palavra_chave.send_keys(palavra_chave)
campo_palavra_chave.send_keys(Keys.RETURN)
time.sleep(333330)

#Posição da Empresa
informacao_da_pagina = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/h3[1]').text
coluna_c = sheet[f'C{linha}']  # Aqui você está definindo a célula correta na coluna C
coluna_c.value = informacao_da_pagina  # Aqui você está atribuindo o valor à célula


# Rolar a página para baixo
#driver.execute_script("window.scrollBy(0, 100);")

# Tirar um print da página
#screenshot_path = "screenshot.png"
#driver.save_screenshot(screenshot_path)

# Criar uma pasta com o nome da empresa (se não existir)
#if not os.path.exists(nome_empresa):
#    os.mkdir(nome_empresa)

# Mover o print para a pasta
#os.rename(screenshot_path, os.path.join(nome_empresa, screenshot_path))

# Renomear o print com o nome da empresa
#nome_arquivo = nome_empresa + ".png"
#os.rename(os.path.join(nome_empresa, screenshot_path), os.path.join(nome_empresa, nome_arquivo))

# Fechar o navegador
#driver.quit()