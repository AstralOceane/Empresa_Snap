import traceback
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook

# Inicia o Chrome
driver = webdriver.Chrome()
driver.get('https://app.localo.com/paywall')
driver.set_window_size(1920, 1080)
sleep(15)

# Digita o email
email = 'matheus.novaes998@gmail.com'
campo_email = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/div/form/div[1]/input')
campo_email.send_keys(email)

# Digita a senha
senha = '321321Asd@'
campo_senha = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/div/form/div[2]/span/input')
campo_senha.send_keys(senha)

# Clicar em conectar
botao_conectar = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div[2]/div/form/button')
botao_conectar.click()
sleep(10)

# Extraia a informação da página usando XPath
informacao_da_pagina = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/h1').text

# Carregue o arquivo Excel existente (dados.xlsx)
workbook = load_workbook('dados.xlsx')

# Selecione a planilha desejada (por exemplo, Planilha1)# Extraia a informação da página usando XPath
informacao_da_pagina = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/h1').text

# Carregue o arquivo Excel existente (dados.xlsx)
workbook = load_workbook('dados.xlsx')

# Selecione a planilha desejada (por exemplo, Planilha1)
sheet = workbook['Plan1']

# Adicione a informação extraída à coluna C a partir da linha 2
linha = 2
coluna = 3  # Coluna C
sheet.cell(row=linha, column=coluna, value=informacao_da_pagina)

# Salve as alterações no arquivo Excel
workbook.save('dados.xlsx')
sheet = workbook['Plan1']

# Adicione a informação extraída à coluna C a partir da linha 2
linha = 2
coluna = 3  # Coluna C
sheet.cell(row=linha, column=coluna, value=informacao_da_pagina)

# Salve as alterações no arquivo Excel
workbook.save('dados.xlsx')

# Feche o WebDriver
driver.quit()
